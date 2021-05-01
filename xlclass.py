"""
# 05.01.2021

Requirements:
python 3.6+

openpyxl==3.0.6
pandas==1.2.2
xlrd==2.0.1

"""

import csv
import datetime
import operator
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Color dict for background fill
COLORS = {'red': PatternFill(fgColor='FF0000', fill_type='solid'),
          'green': PatternFill(fgColor='00b050', fill_type='solid'),
          'orange': PatternFill(fgColor='FFC000', fill_type='solid'),
          'yellow': PatternFill(fgColor='FFFF00', fill_type='solid'),
          'gray': PatternFill(fgColor='C0C0C0', fill_type='solid')}


class Xlsx:
    """
    Class for working with Excel *.xlsx files using Openpyxl.
    Generates an Xlsx object with Openpyxl Workbook/Worksheet objects
    as attributes for use with the enclosed methods.
    """

    def __init__(self, filepath: str = None, sheetname: str = None) -> None:
        """
        Initialize main attributes for Xlsx objects if Path points to
        an existing Excel file. Creates a blank Workbook/Worksheet 
        object if no filepath is passed. If multiple sheets are present 
        in passed Excel file, the name of the sheet you want to work 
        with can be passed as a string to 'sheetname' or you can select 
        needed sheet from a menu. If the Excel file that is passed is an
        *.xls file, sheetname is required and Pandas is used to read the
        sheet data and a new unformatted Xlsx object is created 
        containing that data. 

        Attrs:
            *.path (str/pathlib.Path, optional): Filepath information.
            *.wb (openpyxl.Workbook): Workbook object for Excel file.
            *.ws (openpyxl.Workbook.worksheet): Active sheet for 
                Excel file.

        Args:
            filepath (str/pathlib.Path, optional): str/Path object 
                representing *.xlsx input file.
            sheetname (str, optional): Name representing which sheet you
                want to work with. ex: 'Invoice'
        """
        if filepath:
            # Convert xls to xlsx data using Pandas/Xlrd
            if str(filepath).endswith(".xls"):
                try:
                    # Read data from xls and create xlsx object
                    df = pd.read_excel(filepath, sheet_name=sheetname)
                    self.path = filepath
                    self.wb = openpyxl.Workbook()
                    self.ws = self.wb.active
                    self.ws.title = sheetname
                    # Copy row data from xls to new xlsx object
                    for row in dataframe_to_rows(df):
                        self.ws.append(row)
                    # Remove index row/colum created by Pandas
                    self.ws.delete_cols(1, 1)
                    self.ws.delete_rows(1, 1)

                except Exception as e:
                    print(f"\n Error: {e}\n Error converting from xls be sure "
                          "to include sheetname argument when passing file.")
                    input(" \"sheetname='Invoice'\", etc\n ENTER to close...")
                    exit(" Exiting...")

            else:
                self.path = filepath
                self.wb = openpyxl.load_workbook(filepath)
                # Set first sheet as active if only one is present
                if len(self.wb.sheetnames) == 1:
                    self.ws = self.wb.active

                else:
                    # Set active sheet to sheetname if passed during
                    # object creation
                    if sheetname:
                        self.ws = self.wb[sheetname]

                    else:
                        # Display availible sheets and set worksheet
                        # based on selection if 2+ sheets are present
                        # and sheetname isn't passed.
                        print('\n Which tab/worksheet are we using?\n')
                        for num, sheet in enumerate(self.wb.sheetnames, 1):
                            print(f' {num}: {sheet}')

                        while True:
                            try:
                                self.ws = self.wb[self.wb.sheetnames[int(
                                    input('\n Selection: '))-1]]
                                break
                            except ValueError:
                                print('\n Try again...')

        else:
            # If not file is passed, create a new object and set
            # active worksheet.
            self.path = None
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active

    def copy_sheet_data(self, other: object, columns: dict) -> object:
        """
        Copy cell values from source Excel Worksheet to target Excel 
        Worksheet using a passed dictionary of column letters.

        Args:
            other (Xlsx): Input Xlsx Excel file object to copy 
                values from.
            columns (dict{str: str}): Dictionary of column letters 
                representing the source and target column letters to 
                copy the values. ex: {'A': 'C', 'D': 'B'}

        Returns:
            self: Xlsx object.
        """
        try:
            for row, _cell in enumerate(other.ws['A'], 1):
                for scol, tcol in columns.items():
                    self.ws[f'{tcol.upper()}{row}'] = other.ws[
                        f'{scol.upper()}{row}'].value

        except Exception as e:
            print(f"\nError - copy_sheet_data: {e}")
            input("[ENTER] to continue...")

        return self

    def copy_csv_data(self, incsv: str) -> object:
        """
        Copy all values from csv file to target Excel Worksheet.

        Args:
            incsv (str/pathlib.Path): Path object representing a csv file.

        Returns:
            self: Xlsx object.
        """
        try:
            with open(incsv, 'r') as f:
                reader = csv.reader(f)
                [self.ws.append(row) for row in reader]

        except Exception as e:
            print(f"\nError - copy_csv_data: {e}")
            input("[ENTER] to continue...")

        return self

    def sort_and_replace(self, sortcol: str, startrow: int = 1) -> object:
        """
        Sort and replace cell values based on values of a specific 
        column. Use this BEFORE any cell formatting, etc as it DELETES 
        the values and then replaces them after sorting.

        Args:
            sortcol (str): Column letter containing the values to use as
                "keys" to sort row data by. ex: 'A'
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        try:
            sortme = []
            for row, rowdata in enumerate(self.ws.iter_rows(), 1):
                if row >= startrow:
                    sortme.append([str(self.ws[
                        f'{sortcol.upper()}{row}'].value).lower(), rowdata])

            self.ws.delete_rows(startrow, self.ws.max_row)

            for _sortval, rowdata in sorted(
                    sortme, key=operator.itemgetter(0)):
                self.ws.append(rowdata)

        except Exception as e:
            print(f"\nError - sort_and_replace: {e}")
            input("[ENTER] to continue...")

        return self

    def name_headers(self, headers: dict,
                     hdrrow=1, bold: bool = False) -> object:
        """
        Cycle through header row and fill cells with values.

        Args:
            headers (dict{str:str}): Str pairs of columns and header 
                name values. ex: {'A': 'Name'}
            hdrrow (int, optional): Row to be used for headers. 
                Defaults to 1.
            bold (bool, optional): Option to bold values in headers. 
                Defaults to False.

        Returns:
            self: Xlsx object.
        """
        try:
            for col, name in headers.items():
                self.ws[f'{col.upper()}{hdrrow}'] = name
            if bold:
                for each in self.ws[f'{hdrrow}:{hdrrow}']:
                    each.font = Font(bold=True)

        except Exception as e:
            print(f"\nError - name_headers: {e}")
            input("[ENTER] to continue...")

        return self

    def get_matching_value(self, srchcol: str, srchval: str,
                           retcol: str, startrow: int = 1) -> str:
        """
        Search column for a value and return the corresponding value
        from another column in the same row.

        Args:
            srchcol (str): Column letter to search for a value. ex: 'A'
            srchval (str): Value to search column for. ex: 'Total'
            retcol (str): Column letter containing the corresponding 
                value to be returned. ex: 'B'
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            str: Value from corresponding cell in the same row as search
                value. Returns False if value search value is not found.
        """
        try:
            for row, cell in enumerate(self.ws[srchcol.upper()], 1):
                if row >= startrow and cell.value:
                    if srchval in str(cell.value):
                        return self.ws[f'{retcol.upper()}{row}'].value

        except Exception as e:
            print(f"\nError - get_matching_value: {e}")
            input("[ENTER] to continue...")

        return False

    def set_matching_value(self, srchcol: str, srchval: str, trgtcol: str,
                           setval: str, startrow: int = 1) -> object:
        """
        Search column for a value and set a corresponding value in 
        another column in the same row.

        Args:
            srchcol (str): Column letter to search for a value. ex: 'A'
            srchval (str): Value to search column for. ex: 'Total'
            trgtcol (str): Column letter to set with corresponding 
                value. ex: 'B'
            setval (str): Value to insert into target cell.
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        try:
            for row, cell in enumerate(self.ws[srchcol.upper()], 1):
                if row >= startrow and cell.value:
                    if srchval in str(cell.value):
                        self.ws[f'{trgtcol.upper()}{row}'] = setval

        except Exception as e:
            print(f"\nError - set_matching_value: {e}")
            input("[ENTER] to continue...")

        return self

    def find_remove_row(self, col: str, srch: str, startrow: int = 1) -> object:
        """
        Remove row based on a specific value found in a column.

        Args:
            col (str): Column letter to search for the needed value.
            srch (str): Value to search for.
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        row = startrow
        try:
            while row <= self.ws.max_row:
                if self.ws[f'{col.upper()}{row}'].value:
                    if srch in str(self.ws[f'{col.upper()}{row}'].value):
                        self.ws.delete_rows(row, 1)
                    else:
                        row += 1
                else:
                    row += 1

        except Exception as e:
            print(f"\nError - find_remove_row: {e}")
            input("[ENTER] to continue...")

        return self

    def find_replace(self, col: str, fndrplc: dict,
                     skip: list = None, startrow: int = 1) -> object:
        """
        Search column for a string value and replace it the value is
        not listed in 'skip'.

        Args:
            col (str): Column letter to search for the needed values.
            fndrplc (dict{str: str}): Dictionary of pairs to find and 
                replace. ex: {'find': 'replace'}.
            skip (list(str), optional): List of values to ignore when 
                replacing. Defaults to None.
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        try:
            if not skip:
                skip = []
            for row, cell in enumerate(self.ws[col.upper()], 1):
                if row >= startrow:
                    if cell.value and str(cell.value).lower() not in skip:
                        for find, replace in fndrplc.items():
                            if find in str(cell.value):
                                self.ws[
                                    f'{col.upper()}{row}'
                                ] = str(cell.value).replace(
                                    find, replace)

        except Exception as e:
            print(f"\nError - find_replace: {e}")
            input("[ENTER] to continue...")

        return self

    def move_values(self, scol: str, tcol: str,
                    vals: list, startrow: int = 1) -> object:
        """
        Search source column for passed list of values and 
        move them to target column. 

        Args:
            scol (str): Source column letter to search for values. 
                ex: 'A'
            tcol (str): Target column letter to move located values to. 
                ex: 'B'
            vals (list(str)): List of str values to move. 
                ex: ('name', '20')
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        try:
            for row, cell in enumerate(self.ws[scol.upper()], 1):
                if cell.value and row >= startrow:
                    for item in vals:
                        if item in str(cell.value):
                            self.ws[f'{tcol.upper()}{row}'] = item
                            self.ws[
                                f'{scol.upper()}{row}'] = cell.value.replace(
                                item, '')
                            break

        except Exception as e:
            print(f"\nError - move_values: {e}")
            input("[ENTER] to continue...")

        return self

    def verify_length(self, col: str, length: int, fillcolor: str,
                      skip: list = None, startrow: int = 1,
                      stoprow: int = None) -> object:
        """
        Cycle through values in a column to verify their length marking 
        cells of an incorrect length with a background fill color. 

        Args:
            col (str): Column to search for values. ex: 'B'
            length (int): Total character length for correct values.
            fillcolor (str): Background fill color selection from COLORS
            dict.
            skip (list(str), optional): List of string values to skip 
            when evaluating. Defaults to None.
            startrow (int, optional): Starting row number where values 
            begin. Defaults to 1.
            stoprow (int, optional): Ending row number where values end.
            Defaults to None.

        Returns:
            self: Xlsx object.
        """
        if not stoprow:
            stoprow = self.ws.max_row
        try:
            if not skip:
                skip = []
            if COLORS.get(fillcolor):
                for row, cell in enumerate(self.ws[col.upper()], 1):
                    if startrow <= row <= stoprow:
                        if cell.value and str(cell.value).lower() not in skip:
                            if len(str(cell.value)) != length:
                                self.ws[
                                    f'{col.upper()}{row}'].fill = COLORS.get(
                                    fillcolor)
            else:
                print(f" Color '{fillcolor}' not available.")

        except Exception as e:
            print(f"\nError - verify_length: {e}")
            input("[ENTER] to continue...")

        return self

    def highlight_rows(self, col: str, srch: str,
                       fillcolor: str, startrow: int = 1) -> object:
        """
        Search row for specified str value and fill entire row 
        with specified background fill color when found. 

        Args:
            col (str): Column to search for value. ex: 'B'
            srch (str): Str value to search cells for.
            fillcolor (str): Background fill color selection from COLORS 
                dict.
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        try:
            if COLORS.get(fillcolor):
                for row, cell in enumerate(self.ws[col.upper()], 1):
                    if row >= startrow:
                        if cell.value and srch.lower() in str(
                                cell.value).lower():
                            for each in self.ws[f'{row}:{row}']:
                                each.fill = COLORS.get(fillcolor)
            else:
                print(f" Color '{fillcolor}' not available.")

        except Exception as e:
            print(f"\nError - highlight_rows: {e}")
            input("[ENTER] to continue...")

        return self

    def number_type_fix(self, col: str,
                        numtype: str, startrow: int = 1) -> object:
        """
        Quick fix for cells that contain numbers formatted as 
        text/str data. Cycle through cells replacing str formatted 
        values with int/float values.

        Args:
            col (str): Column containing data to convert
            numtype (str): 'i' or 'f' indicating which type of number 
                values the column contains (int/float)
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        try:
            for row, cell in enumerate(self.ws[col.upper()], 1):
                if cell.value and row >= startrow:
                    if numtype.lower() == 'i':
                        self.ws[f'{col.upper()}{row}'] = int(cell.value)
                    if numtype.lower() == 'f':
                        self.ws[f'{col.upper()}{row}'] = float(cell.value)

        except Exception as e:
            print(f"\nError - number_type_fix: {e}")
            input("[ENTER] to continue...")

        return self

    def format_date(self, col: str, startrow: int = 1) -> object:
        """
        Format str date value to (MM/DD/YYYY).

        Args:
            col (str): Column containing date values.
            startrow (int, optional): Starting row number where values 
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        try:
            for row, cell in enumerate(self.ws[col.upper()], 1):
                if row >= startrow and cell.value:
                    self.ws[f'{col.upper()}{row}'] = cell.value.strftime(
                        '%m/%d/%Y')

        except Exception as e:
            print(f"\nError - format_date: {e}")
            input("[ENTER] to continue...")

        return self

    def format_currency(self, col: str,
                        startrow: int = 1, stoprow: int = None) -> object:
        """
        Format str currency value to ($0,000.00).

        Args:
            col (str): Column containing currency values to be formatted.
            startrow (int, optional): Starting row number where values 
            begin. Defaults to 1.
            stoprow (int, optional): Ending row where values stop.
            Defaults to None.

        Returns:
            self: Xlsx object.
        """
        if not stoprow:
            stoprow = self.ws.max_row
        try:
            for row, cell in enumerate(self.ws[col.upper()], 1):
                if startrow <= row <= stoprow and cell.value:
                    cell.number_format = '$#,###.00'

        except Exception as e:
            print(f"\nError - format_currency: {e}")
            input("[ENTER] to continue...")

        return self

    def set_cell_size(self, pairs: dict) -> object:
        """
        Selects rows and columns and adjusts their sizes using a 
        dictionary of pairs of rows or columns along with corresponding 
        height or width to adjust the size of cells from each pair. If 
        dict key is type: str, adjusts column width. If dict key is 
        type: int, adjusts row height. 

        Args:
            pairs (dict): Dictionary of column/row keys with target size  
                as their values.

        Returns:
            self: Xlsx object.
        """
        try:
            for target, size in pairs.items():
                if type(target) == str:
                    self.ws.column_dimensions[target.upper()].width = size
                elif type(target) == int:
                    self.ws.row_dimensions[target].height = size
                else:
                    print(
                        f"Invalid data pair. Check your info. {target: size}")
                    input("[ENTER] to continue...")

        except Exception as e:
            print(f"\nError - set_cell_size: {e}")
            input("[ENTER] to continue...")

        return self

    def save(self, savepath: str = None) -> None:
        """
        Duplicates openpyxl's save function so it can be called on the 
        object without needing the .wb attribute, etc. Saves the Excel 
        file to the specified filepath or Path location if passed. If no 
        filepath is passed, uses the original file's Path (.path attr) 
        to save over the original.

        Args:
            savepath (str or pathlib.Path, optional): Output file 
                location (including filename) for your output file. Uses
                original if not specified. Defaults to None.
        """
        try:
            if savepath:
                self.wb.save(savepath)
            elif self.path:
                self.wb.save(self.path)
            else:
                input("\n No savepath found...")

        except Exception as e:
            print(f"\nError - save: {e}")
            input("[ENTER] to continue...")

    def generate_dictionary(self, datacols: list, keycol: str = None,
                            hdrrow: int = 1, datastartrow: int = None) -> dict:
        """
        Reads the headers and cells from the spreadsheet and usees them 
        to generate a dictionary of the data. Data listed in *keycol* on 
        spreadsheet will need to be a series of unique values to be used 
        as keys or the information assigned will be overwritten each 
        time a duplicate key is found. If *keycol* is not specified, a 
        4-digit string of the row number is used for each key. 
        ex: '0005'

        Args:
            datacols (list): List of string column letters where needed 
                data is located.
            keycol (str, optional): Column letter where the data that 
                will be used as the dictionary keys is located. If not 
                passed, 4-digit string of the row numbers will be used 
                instead. Defaults to None.
            hdrrow (int, optional) Row number containing the headers in 
                the spreadsheet. Defaults to 1.
            datastartrow (int, optional) Row number where the needed 
                data starts. If not specified, data will be read from 
                header row + 1. Defaults to None.

        Returns:
            dict: Dictionary generated from the data in the spreadsheet. 
                {key: {header: value}}
        """
        data = {}
        keycolumn = keycol if keycol else 'A'
        datastart = hdrrow + 1 if not datastartrow else datastartrow
        try:
            for row, cell in enumerate(self.ws[keycolumn.upper()], 1):
                keys = cell.value if keycol else f"{row:0>4}"
                if row >= datastart and keys:
                    data[keys] = {
                        self.ws[f'{ea.upper()}{hdrrow}'].value: self.ws[
                            f'{ea.upper()}{row}'].value for ea in datacols}

            return data

        except Exception as e:
            print(f"\nError - generate_dictionary: {e}")
            input("[ENTER] to continue...")

    def generate_list(self, startrow: int = 1, stoprow: int = None) -> list:
        """
        Generates a list of lists containing all cell values from 
        startrow to stoprow (inclusive). (Use _list.pop(0) on returned 
        list to get a separate headers list if present/needed.)

        Args:
            startrow (int, optional): First row to pull data from to 
                generate the list. Defaults to 1.
            stoprow (int, optional): Last row to pull data from to 
                generate the list. If no value is passed, it will pull 
                data from all rows after startrow. Defaults to None.

        Returns:
            list: List of lists containing the values read from cells.
        """
        row_data = []
        for row in self.ws.iter_rows(min_row=startrow, max_row=stoprow):
            row_data.append([cell.value for cell in row])

        return row_data
