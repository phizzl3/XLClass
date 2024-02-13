import csv
import datetime
import operator
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

from .utils import (
    _convert_xls,
    _generate_source_target_columns_dictionary,
    generate_columns_dictionary,
)

from .write_dictionary_to_sheet import _write_dictionary_to_sheet

# Color dict for background fill
COLORS = {
    "red": PatternFill(fgColor="FF0000", fill_type="solid"),
    "green": PatternFill(fgColor="00b050", fill_type="solid"),
    "orange": PatternFill(fgColor="FFC000", fill_type="solid"),
    "yellow": PatternFill(fgColor="FFFF00", fill_type="solid"),
    "gray": PatternFill(fgColor="C0C0C0", fill_type="solid"),
}


class Xlsx:
    """Class for working with Excel *.xlsx files using Openpyxl.
    Generates an Xlsx object with Openpyxl Workbook/Worksheet objects
    as attributes for use with the enclosed methods.
    """

    def __init__(self, filepath: str = None, sheetname: str = None) -> None:
        """Initialize main attributes for Xlsx objects if Path points to
        an existing Excel file. Creates a blank Workbook/Worksheet
        object if no filepath is passed. If multiple sheets are present
        in passed Excel file, the name of the sheet you want to work
        with can be passed as a string to 'sheetname' or you can select
        needed sheet from a menu. If the Excel file that is passed is an
        *.xls file, sheetname is required and Pandas is used to read the
        sheet data and a new unformatted Xlsx object is created
        containing that data.

        Attrs:
            *.path (pathlib.Path, optional): Filepath information.
            Defults to None if new blank object is created.
            *.wb (openpyxl.Workbook): Workbook object for Excel file.
            *.ws (openpyxl.Workbook.worksheet): Active sheet for
            Excel file.

        Args:
            filepath (str/pathlib.Path, optional): str/Path object
            representing *.xlsx input file.
            sheetname (str, optional): Name representing which sheet you
            want to work with. ex: 'Invoice'
        """
        if filepath:
            # Convert xls to xlsx data using Pandas/Xlrd
            if str(filepath).endswith(".xls"):
                _convert_xls(self, filepath, sheetname)

            elif str(filepath).endswith(".xlsx"):
                self.path = Path(filepath)
                self.wb = openpyxl.load_workbook(filepath)

                # Set first sheet as active if only one is present
                if len(self.wb.sheetnames) == 1:
                    self.ws = self.wb.active

                else:
                    # Set active sheet to sheetname if passed during
                    # object creation
                    if sheetname:
                        self.ws = self.wb[sheetname]

                    else:
                        # Display availible sheets and set worksheet
                        # based on selection if 2+ sheets are present
                        # and sheetname isn't passed.
                        print("\n Which tab/worksheet are we using?\n")
                        for num, sheet in enumerate(self.wb.sheetnames, 1):
                            print(f" {num}: {sheet}")

                        while True:
                            try:
                                self.ws = self.wb[
                                    self.wb.sheetnames[int(input("\n Selection: ")) - 1]
                                ]
                                break
                            except ValueError:
                                print("\n Try again...")

            else:
                input("File not supported. Please use .xlsx or xls.")
                exit("Exiting...")

        else:
            # If not file is passed, create a new object and set
            # active worksheet.
            self.path = None
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active

    def save(self, savepath: str = None) -> None:
        """Duplicates openpyxl's save function so it can be called on the
        object without needing the .wb attribute, etc. Saves the Excel
        file to the specified filepath or Path location if passed. If no
        filepath is passed, uses the original file's Path (.path attr)
        to save over the original.

        Args:
            savepath (str or pathlib.Path, optional): Output file
                location (including filename) for your output file. Uses
                original if not specified. Defaults to None.
        """
        if savepath:
            self.wb.save(savepath)
        elif self.path:
            self.wb.save(self.path)
        else:
            input("\n No savepath found...")

    def generate_headers_attribute(self, header_row: int = 1):
        """Uses specified header row number to generate a *.headers
        attribute containing a dictionary of header values and their
        corresonding column letters. {"Header 1": "A", "Header 2": "B"}

        Args:
            header_row (int, optional): Row containing header values.
            Defaults to 1.

        Returns:
            self: Xlsx object (that includes *.headers attribute)
        """
        # Skip to specified header row, stop at next row
        for row, row_data in enumerate(self.ws.iter_rows(), 1):
            if row < header_row:
                continue
            if row == header_row + 1:
                break
            # Generate a list of header values
            header_values = [cell.value for cell in row_data]
        # Generate the dictionary and add to *.headers attribute
        self.headers = generate_columns_dictionary(key_list=header_values)

        return self

    def copy_sheet_data(self, source: object, columns: dict):
        """Copy cell values from source Excel Worksheet to target (self)
        Excel Worksheet using a passed dictionary of column letters.

        Args:
            source (Xlsx): Input Xlsx Excel file object to copy values from.
            columns (dict{str: str}): Dictionary of column letters
                representing the source and target column letters to
                copy the values. ex: {'A': 'C', 'D': 'B'}

        Returns:
            self: Xlsx object.
        """
        for row, _cell in enumerate(source.ws["A"], 1):
            for scol, tcol in columns.items():
                self.ws[f"{tcol.upper()}{row}"] = source.ws[
                    f"{scol.upper()}{row}"
                ].value

        return self

    def copy_sheet_data_by_headers(
        self, source: object, source_dict: dict, keep_list: list
    ):
        """Copy cell values from source Excel Worksheet to target (self)
        Excel worksheet using a passed dictionary of {header: column}
        matching the source worksheet, and a list of headers corresponding
        to the columns that need to be copied.

        Args:
            source (Xlsx): Input Xlsx Excel file object to copy values from.
            source_dict (dict): Dictionary of header values and corresponding
            columns matching the source Excel file object.
            keep_list (list): Ordered list of strings matching the header
            values that you want to keep/copy into the target file object.

        Returns:
            self: Xlsx object.
        """
        # Generate {source: target} columns dictionary
        source_target = _generate_source_target_columns_dictionary(
            source_dict=source_dict, keep_list=keep_list
        )
        # Copy sheet data using columns generated above
        self.copy_sheet_data(source=source, columns=source_target)

        return self

    def copy_csv_data(self, source_csv: str):
        """Copy all values from source csv file to target Excel Worksheet.

        Args:
            source_csv (str/pathlib.Path): Path object representing a csv file.

        Returns:
            self: Xlsx object.
        """
        with open(source_csv, "r") as f:
            reader = csv.reader(f)
            [self.ws.append(row) for row in reader]

        return self

    def sort_and_replace(self, sortcol: str, startrow: int = 1):
        """Sort and replace cell values based on values of a specific
        column. Use this BEFORE any cell formatting, etc as it DELETES
        the values and then replaces them after sorting.

        Args:
            sortcol (str): Column letter containing the values to use as
                "keys" to sort row data by. ex: 'A'
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        sortme = []
        for row, rowdata in enumerate(self.ws.iter_rows(), 1):
            if row >= startrow:
                sortme.append(
                    [str(self.ws[f"{sortcol.upper()}{row}"].value).lower(), rowdata]
                )

        self.ws.delete_rows(startrow, self.ws.max_row)

        for _sortval, rowdata in sorted(sortme, key=operator.itemgetter(0)):
            self.ws.append(rowdata)

        return self

    def name_headers(self, headers: dict, hdrrow=1, bold: bool = False):
        """Cycle through header row and fill cells with values.

        Args:
            headers (dict{str:str}): Str pairs of columns and header
                name values. ex: {'A': 'Name'}
            hdrrow (int, optional): Row to be used for headers.
                Defaults to 1.
            bold (bool, optional): Option to bold values in headers.
                Defaults to False.

        Returns:
            self: Xlsx object.
        """
        for col, name in headers.items():
            self.ws[f"{col.upper()}{hdrrow}"] = name
        if bold:
            for each in self.ws[f"{hdrrow}:{hdrrow}"]:
                each.font = Font(bold=True)

        return self

    def set_matching_value(
        self, srchcol: str, srchval: str, trgtcol: str, setval: str, startrow: int = 1
    ):
        """Search column for a value and set a corresponding value in
        another column in the same row.

        Args:
            srchcol (str): Column letter to search for a value. ex: 'A'
            srchval (str): Value to search column for. ex: 'Total'
            trgtcol (str): Column letter to set with corresponding
                value. ex: 'B'
            setval (str): Value to insert into target cell.
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        for row, cell in enumerate(self.ws[srchcol.upper()], 1):
            if row >= startrow and cell.value:
                if srchval in str(cell.value):
                    self.ws[f"{trgtcol.upper()}{row}"] = setval

        return self

    def find_remove_row(self, col: str, srch: str, startrow: int = 1):
        """Remove row based on a specific value found in a column.

        Args:
            col (str): Column letter to search for the needed value.
            srch (str): Value to search for.
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        row = startrow
        while row <= self.ws.max_row:
            if self.ws[f"{col.upper()}{row}"].value:
                if srch in str(self.ws[f"{col.upper()}{row}"].value):
                    self.ws.delete_rows(row, 1)
                else:
                    row += 1
            else:
                row += 1

        return self

    def find_replace(
        self, col: str, fndrplc: dict, skip: list = None, startrow: int = 1
    ):
        """Search column for a string value and replace it the value is
        not listed in 'skip'.

        Args:
            col (str): Column letter to search for the needed values.
            fndrplc (dict{str: str}): Dictionary of pairs to find and
                replace. ex: {'find': 'replace'}.
            skip (list(str), optional): List of values to ignore when
                replacing. Defaults to None.
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        if not skip:
            skip = []
        for row, cell in enumerate(self.ws[col.upper()], 1):
            if row >= startrow:
                if cell.value and str(cell.value).lower() not in skip:
                    for find, replace in fndrplc.items():
                        if find in str(cell.value):
                            self.ws[f"{col.upper()}{row}"] = str(cell.value).replace(
                                find, replace
                            )

        return self

    def move_values(self, scol: str, tcol: str, vals: list, startrow: int = 1):
        """Search source column for passed list of values and
        move them to target column.

        Args:
            scol (str): Source column letter to search for values.
                ex: 'A'
            tcol (str): Target column letter to move located values to.
                ex: 'B'
            vals (list(str)): List of str values to move.
                ex: ('name', '20')
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        for row, cell in enumerate(self.ws[scol.upper()], 1):
            if cell.value and row >= startrow:
                for item in vals:
                    if item in str(cell.value):
                        self.ws[f"{tcol.upper()}{row}"] = item
                        self.ws[f"{scol.upper()}{row}"] = cell.value.replace(item, "")
                        break

        return self

    def reverse_text(self, datacol: str = "A", startrow: int = 1, separator: str = ","):
        """Get values from specified column, split them on specified separator,
        reverse the value's order and write them back to the cell minus the
        separator. ex: Last, First -> First Last


        Args:
            datacol (str, optional): Excel column with values.
            Defaults to "A".
            startrow (int, optional): Excel row where values begin.
            Defaults to 1.
            separator (str, optional): Text separator to split on.
            Defaults to ",".

        Returns:
            self: Xlsx object
        """
        for row, cell in enumerate(self.ws[datacol.upper()], 1):
            if row < startrow or not cell.value or separator not in cell.value:
                continue

            # Swap info and write back to cell
            split_value = str(cell.value).split(separator)

            self.ws[f"{datacol.upper()}{row}"] = (
                f"{split_value[1].strip()} {split_value[0].strip()}"
            )

        return self

    def remove_non_numbers(
        self, datacol: str, startrow: int = 1, stoprow: int = None, skip: list = []
    ):
        """Get values from a specified column that should contain only
        numbers. Remove any characters that are non-numbers and write
        the new values back to the cells (as a string value). If a list is
        passed to skip, check this list first before processing and skip
        the cell if it matches an entry in the list.

        Args:
            datacol (str): Excel column with values to clean.
            startrow (int): Excel row where values begin.
            stoprow (int, optional): Excel row to stop cleaning values.
            Defaults to None.
            skip (list, optional): List of string values to skip if
            found in the specified cells. Defaults to an empty list.

        Returns:
            self: Xlsx object
        """
        for row, cell in enumerate(self.ws[f"{datacol.upper()}"], 1):
            if stoprow and row == stoprow:
                break
            if row < startrow or not cell.value:
                continue
            if str(cell.value).lower() in skip:
                continue
            # Read and clean the data leaving only numbers
            new_value = cell.value
            for char in str(cell.value):
                if char.isnumeric():
                    continue
                new_value = str(cell.value).replace(char, "")
            # Replace cell value with new version
            self.ws[f"{datacol.upper()}{row}"] = new_value

        return self

    def get_matching_value(
        self, srchcol: str, srchval: str, retcol: str, startrow: int = 1
    ) -> str:
        """Search column for a value and return the corresponding value
        from another column in the same row.

        Args:
            srchcol (str): Column letter to search for a value. ex: 'A'
            srchval (str): Value to search column for. ex: 'Total'
            retcol (str): Column letter containing the corresponding
                value to be returned. ex: 'B'
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            str: Value from corresponding cell in the same row as search
                value. Returns False if value search value is not found.
        """
        for row, cell in enumerate(self.ws[srchcol.upper()], 1):
            if row >= startrow and cell.value:
                if srchval in str(cell.value):
                    return self.ws[f"{retcol.upper()}{row}"].value

        return False

    def search_matching_value(self, header_srch_value: str, row_srch_value: str) -> str:
        """Searches cells by row for header search value and row search
        value, and returns corresponding cell value matching both as
        a string.

        Args:
            header_srch_value (str): Header name to search for.
            row_srch_value (str): Row name to search for.

        Returns:
            str: Matching (intersecting) value corresponding to the
            searched header and row value. Returns False if not found.
        """
        search_column, search_row = 0, 0

        for row in self.ws.iter_rows():
            for cell_number, cell_data in enumerate(row, 1):
                if str(cell_data.value) == header_srch_value:
                    search_column += cell_number

                if str(cell_data.value) == row_srch_value:
                    search_row = True

                if search_row:
                    if cell_number == search_column:
                        return str(cell_data.value)

        # In case search isn't located.
        return False

    def verify_length(
        self,
        col: str,
        length: int,
        fillcolor: str,
        skip: list = None,
        startrow: int = 1,
        stoprow: int = None,
    ):
        """Cycle through values in a column to verify their length marking
        cells of an incorrect length with a background fill color.

        Args:
            col (str): Column to search for values. ex: 'B'
            length (int): Total character length for correct values.
            fillcolor (str): Background fill color selection from COLORS
            dict.
            skip (list(str), optional): List of string values to skip
            when evaluating. Defaults to None.
            startrow (int, optional): Starting row number where values
            begin. Defaults to 1.
            stoprow (int, optional): Ending row number where values end.
            Defaults to None.

        Returns:
            self: Xlsx object.
        """
        if not stoprow:
            stoprow = self.ws.max_row
        if not skip:
            skip = []
        if COLORS.get(fillcolor.lower()):
            for row, cell in enumerate(self.ws[col.upper()], 1):
                if startrow <= row <= stoprow:
                    if cell.value and str(cell.value).lower() not in skip:
                        if len(str(cell.value)) != length:
                            self.ws[f"{col.upper()}{row}"].fill = COLORS.get(
                                fillcolor.lower()
                            )
        else:
            print(f" Color '{fillcolor}' not available.")

        return self

    def find_and_highlight_rows(
        self, col: str, srch: str, fillcolor: str = "red", startrow: int = 1
    ):
        """Search row for specified str value and fill entire row
        with specified background fill color when found.

        Args:
            col (str): Column to search for value. ex: 'B'
            srch (str): Str value to search cells for.
            fillcolor (str): Background fill color selection from COLORS
                dict.
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        if COLORS.get(fillcolor.lower()):
            for row, cell in enumerate(self.ws[col.upper()], 1):
                if row >= startrow:
                    if cell.value and srch.lower() in str(cell.value).lower():
                        for each in self.ws[f"{row}:{row}"]:
                            each.fill = COLORS.get(fillcolor.lower())
        else:
            print(f" Color '{fillcolor}' not available.")

        return self

    def number_type_fix(self, col: str, numtype: str, startrow: int = 1):
        """Quick fix for cells that contain numbers formatted as
        text/str data. Cycle through cells replacing str formatted
        values with int/float values.

        Args:
            col (str): Column containing data to convert
            numtype (str): 'i' or 'f' indicating which type of number
                values the column contains (int/float)
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        for row, cell in enumerate(self.ws[col.upper()], 1):
            if cell.value and row >= startrow:
                if numtype.lower() == "i":
                    self.ws[f"{col.upper()}{row}"] = int(cell.value)
                if numtype.lower() == "f":
                    self.ws[f"{col.upper()}{row}"] = float(cell.value)

        return self

    def format_date(self, col: str, startrow: int = 1):
        """Format str date value to (MM/DD/YYYY).

        Args:
            col (str): Column containing date values.
            startrow (int, optional): Starting row number where values
                begin. Defaults to 1.

        Returns:
            self: Xlsx object.
        """
        for row, cell in enumerate(self.ws[col.upper()], 1):
            if row >= startrow and cell.value:
                self.ws[f"{col.upper()}{row}"] = cell.value.strftime("%m/%d/%Y")

        return self

    def format_currency(self, col: str, startrow: int = 1, stoprow: int = None):
        """Format str currency value to ($0,000.00).

        Args:
            col (str): Column containing currency values to be formatted.
            startrow (int, optional): Starting row number where values
            begin. Defaults to 1.
            stoprow (int, optional): Ending row where values stop.
            Defaults to None.

        Returns:
            self: Xlsx object.
        """
        if not stoprow:
            stoprow = self.ws.max_row

        for row, cell in enumerate(self.ws[col.upper()], 1):
            if startrow <= row <= stoprow and cell.value:
                cell.number_format = "$#,###.00"

        return self

    def set_cell_size(self, pairs: dict):
        """Selects rows and columns and adjusts their sizes using a
        dictionary of pairs of rows or columns along with corresponding
        height or width to adjust the size of cells from each pair. If
        dict key is type: str, adjusts column width. If dict key is
        type: int, adjusts row height.

        Args:
            pairs (dict): Dictionary of column/row keys with target size
                as their values.

        Returns:
            self: Xlsx object.
        """
        for target, size in pairs.items():
            if type(target) == str:
                self.ws.column_dimensions[target.upper()].width = size
            elif type(target) == int:
                self.ws.row_dimensions[target].height = size
            else:
                print(f"Invalid data pair. Check your info. {target: size}")
                input("[ENTER] to continue...")

        return self

    def set_bold_rows(self, startrow: int = 1, stoprow: int = 0):
        """Sets all cells in specified rows to bold beginning at startrow
        and ending just before stoprow (if passed). Sets all cells below
        startrow to bold if stoprow isn't passed. *Uses the default font
        family and size settings and overrides any other styles.

        Args:
            startrow (int, optional): Row number where bold text should begin.
            Defaults to 1.
            stoprow (int, optional): Row number (not included) where bold text
            should stop. Defaults to 0.

        Returns:
            self: Xlsx object.
        """
        for row_number, row in enumerate(self.ws.iter_rows(), 1):
            if row_number < startrow:
                continue
            if stoprow and row_number == stoprow:
                break
            for cell in row:
                cell.font = Font(bold=True)

        return self

    def highlight_rows(
        self,
        startrow: int = 1,
        stoprow: int = 0,
        fillcolor: str = "gray",
        alternate: bool = False,
    ):
        """Highlights specified rows (optionally alternating) using passed
        color (from xlclass.COLORS dict) starting at startrow and ending
        just before stoprow. Highlights all remaining rows if stoprow is
        not passed.

        Args:
            startrow (int, optional): Row number where highlighting should
            begin. Defaults to 1.
            stoprow (int, optional): Row number (not included in highlights)
            where highlighting should end. Defaults to 0.
            fillcolor (str, optional): Color choice from xlclass.COLORS
            dictionary to be used as fill color. Defaults to 'gray'.
            alternate (bool, optional): Option to alternate rows to
            highlight. Defaults to False.

        Returns:
            self: Xlsx object.
        """
        if not COLORS.get(fillcolor.lower()):
            print(f"Color: '{fillcolor}' not available.")
            return self

        highlight_row = startrow
        for row_number, row in enumerate(self.ws.iter_rows(), 1):
            if row_number < startrow:
                continue
            if row_number == stoprow:
                break
            if row_number == highlight_row:
                for cell in row:
                    cell.fill = COLORS.get(fillcolor.lower())
                if not alternate:
                    highlight_row += 1
                else:
                    highlight_row += 2

        return self

    def set_sheet_font_style(self, fontname: str = "Arial", size: int = 8):
        """Sets all cells to specified font name and size.
        *Overrides any other font style settings in selected cells.

        Args:
            fontname (str, optional): Font name to use. Defaults to 'Arial'.
            size (int, optional): Font size to use. Defaults to 8.

        Returns:
            self: Xlsx object.
        """
        for row in self.ws.iter_rows():
            for cell in row:
                cell.font = Font(name=fontname, size=str(size))

        return self

    def add_cell_borders(self, startrow: int = 1, stoprow: int = 0):
        """Set thin cell borders around all populated cells beginning at
        startrow and ending at stoprow. If no stoprow is passed, borders
        will be added until the end of the populated cells.

        Args:
            startrow (int, optional): Row number where borders should begin.
            Defaults to 1.
            stoprow (int, optional): Row number where borders should end.
            Defaults to 0.

        Returns:
            self: Xlsx object
        """
        for row_num, row_data in enumerate(self.ws.iter_rows(), 1):
            if row_num < startrow:
                continue
            if stoprow and row_num == stoprow:
                break
            for cell in row_data:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        return self

    def generate_dictionary(
        self,
        datacols: list = [],
        keycol: str = "",
        hdrrow: int = 1,
        datastartrow: int = 2,
    ) -> dict:
        """Reads the headers and cells from the spreadsheet and uses them
        to generate a dictionary of the data. Data listed in *keycol* on
        spreadsheet will need to be a series of unique values to be used
        as keys or the information assigned will be overwritten each
        time a duplicate key is found. If *keycol* is not specified, a
        4-digit string of the row number is used for each key.
        ex: '0005'

        Args:
            datacols (list, optional): List of string column letters where needed
                data is located eg. ["A", "B", "C"] If none is specified,
                reads the entire spreadsheet. Defaults to None.
            keycol (str, optional): Column letter where the data that
                will be used as the dictionary keys is located. If not
                passed, 4-digit string of the row numbers will be used
                instead. Defaults to None.
            hdrrow (int, optional) Row number containing the headers in
                the spreadsheet. Defaults to 1.
            datastartrow (int, optional) Row number where the needed
                data starts. Defaults to 2.

        Returns:
            dict: Nested dictionary generated from the data in the spreadsheet.
                {key: {header: value}}
        """
        data = {}
        keycolumn = keycol if keycol else "A"

        # Runs if data columns are specified and only gets the data in those listed
        if datacols:
            for row, cell in enumerate(self.ws[keycolumn.upper()], 1):
                # Use the value in the key column if specified
                # Otherwise, generates a number based on enumerate/row number
                keys = cell.value if keycol else f"{row:0>4}"
                if row >= datastartrow and keys:
                    data[keys] = {
                        self.ws[f"{ea.upper()}{hdrrow}"]
                        .value: self.ws[f"{ea.upper()}{row}"]
                        .value
                        for ea in datacols
                    }

        # Runs if no data columns are specified and gets the entire sheet
        else:
            headers_list = []
            for row, row_data in enumerate(self.ws.iter_rows(), 1):
                # Use the value in the key column if specified
                # Otherwise, generates a
                keys = self.ws[f"{keycol}{row}"].value if keycol else f"{row:0>4}"
                # Generates a list of values to use as headers (indexed below)
                if row == hdrrow:
                    for cell in row_data:
                        headers_list.append(cell.value)
                if row >= datastartrow:
                    data[keys] = {
                        headers_list[index]: cell.value
                        for index, cell in enumerate(row_data)
                    }

        return data

    def generate_list(self, startrow: int = 1, stoprow: int = None) -> list:
        """Generates a list of lists containing all cell values from
        startrow to stoprow (inclusive). (Use _list.pop(0) on returned
        list to get a separate headers list if present/needed.)

        Args:
            startrow (int, optional): First row to pull data from to
                generate the list. Defaults to 1.
            stoprow (int, optional): Last row to pull data from to
                generate the list. If no value is passed, it will pull
                data from all rows after startrow. Defaults to None.

        Returns:
            list: List of lists containing the values read from cells.
        """
        row_data = []
        for row in self.ws.iter_rows(min_row=startrow, max_row=stoprow):
            row_data.append([cell.value for cell in row])

        return row_data

    def write_dictionary_to_sheet(
        self, data_dict: dict, header_row: int = None, start_row: int = None
    ):
        """Receives a nested dictionary:
        {"Row1": {"Key/Header": Value, "Key/Header": Value,},}
        Writes the headers and row data to the Xlsx object.
        If no header_row is passed, the default will be row 1
        and the default start_row will be the header_row+1.

            Args:
                xlsx (Xlsx): Xlsx object to write data to.
                data_dict (dict): Nested dictionary with data to write to Xlsx.
                header_row (int, optional): Row number to write headers. Defaults to None.
                start_row (int, optional): Row number to start writing data. Defaults to None.
        """
        _write_dictionary_to_sheet(
            xlsx=self, data_dict=data_dict, header_row=header_row, start_row=start_row
        )
        return self
