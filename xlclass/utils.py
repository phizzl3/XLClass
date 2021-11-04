from pathlib import Path

try:
    import pandas as pd
except ImportError:
    pd = False

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def _convert_xls(obj, filepath=None, sheetname=None):
    """Converts .xls data to Xlsx object."""
    if not pd:
        input(".xls support requirements missing. Check requirements.txt")
        exit("Exiting...")

    # Convert xls to xlsx data using Pandas/Xlrd
    if not filepath and sheetname:
        input("Error converting from xls.\nBe sure to include the sheetname "
              "when passing your filepath.")
        exit("Exiting...")

    # Read data from xls and create xlsx object
    df = pd.read_excel(filepath, sheet_name=sheetname)
    obj.path = Path(filepath)
    obj.wb = openpyxl.Workbook()
    obj.ws = obj.wb.active
    obj.ws.title = sheetname

    # Copy row data from xls to new xlsx object
    for row in dataframe_to_rows(df):
        obj.ws.append(row)

    # Remove index row/colum created by Pandas
    obj.ws.delete_cols(1, 1)
    obj.ws.delete_rows(1, 1)


def generate_columns_dictionary(key_list: list) -> dict:
    """Uses the passed ordered list (key_list) of values to generate a
    dictionary of corresponding column letters.
    {list0: "A", list1: "B", list2: "C"}

    Args:
        key_list (list): Ordered list of values to be used as the keys
        in the generated dictionary.

    Returns:
        dict: Dictionary of list items and their corresponding column
        letters. {list0: "A", list1: "B", list2: "C"}
    """
    return {key_value: get_column_letter(
        column_number) for column_number, key_value in enumerate(key_list, 1)}


def _generate_source_target_columns_dictionary(
        source_dict: dict, keep_list: list) -> dict:
    """Uses the passed ordered list (keep_list) to generate a new 
    dictionary from matching values using the keys from the passed 
    dictionary (source_dict) to generate a new dictionary with 
    corresponding target column numbers in the order of keep_list.
    ex: {(1st matching value from source_dict) "B": "A"}

    Args:
        source_dict (dict): Dictionary of headers and matching columns
        from source Xlsx object.
        keep_list (list): Ordered list of (exact) string values of header
        values that contain the data you'd like to keep.

    Returns:
        dict: Dictionary of source: target letters {"B": "A"}
    """
    return {source_dict[keep_value]: get_column_letter(
        column_number) for column_number, keep_value in enumerate(
            keep_list, 1)}
