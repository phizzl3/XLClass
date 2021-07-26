"""
Run from top-level folder as module: 

$ python -m tests (-v for verbose mode)

Requirements:
python 3.6+

openpyxl==3.0.6
pandas==1.2.2
xlrd==2.0.1

"""

import datetime
import unittest
from pathlib import Path

import openpyxl
from xlclass import Xlsx

tests_path = Path(__file__).resolve().parent
test_xlsx = tests_path / "_test.xlsx"
test_csv = tests_path / "_test.csv"
test_xls = tests_path / "_test.xls"


class TestXlsx(unittest.TestCase):

    def setUp(self):
        """
        Generate test Xlsx object for each test. 
        """
        self.xl = Xlsx(test_xlsx)

    def test_init(self):
        """
        Test attributes created on __init__() are generated as expected.
        Checks .path ends in correct filename. Checks .wb and .ws are 
        instances of the correct openpyxl objects.
        """
        self.assertTrue(str(self.xl.path).endswith("_test.xlsx"))
        self.assertIsInstance(self.xl.wb, openpyxl.Workbook)
        self.assertIsInstance(
            self.xl.ws, openpyxl.worksheet.worksheet.Worksheet)

    def test_xls_conversion(self):
        """
        Test xls -> xlsx conversion during __init__() and verify cell 
        data returns as expected.
        """
        self.xl = Xlsx(test_xls, 'Sheet1')
        self.assertIsInstance(self.xl.wb, openpyxl.Workbook)
        self.assertEqual(self.xl.ws['A9'].value, 'H')
        self.assertEqual(self.xl.ws['B12'].value, 'Triangle')
        self.assertEqual(self.xl.ws['E20'].value, 433.0498)
        self.iterate_integers()

    def test_copy_sheet_data(self):
        """
        Create a temporary blank workbook, copy selected data to the 
        workbook, and test cell data returns as expected.
        """
        self.xl_temp = Xlsx()
        self.xl_temp.copy_sheet_data(self.xl, {'B': 'A', 'D': 'B', 'C': 'C'})
        self.assertEqual(self.xl_temp.ws['A9'].value, 'Magenta')
        self.assertEqual(self.xl_temp.ws['C20'].value, 1900)
        self.iterate_integers()

    def test_copy_csv_data(self):
        """
        Delete all data from main test workbook, open test csv file and
        copy all data to main workbook, and test cell data returns as 
        expected.
        """
        self.xl.ws.delete_rows(1, self.xl.ws.max_row)
        self.xl.copy_csv_data(test_csv)
        self.assertEqual(self.xl.ws['A1'].value, 'State')
        self.assertEqual(self.xl.ws['B1'].value, 'Number')
        self.assertEqual(self.xl.ws['A11'].value, 'Georgia')
        self.assertEqual(self.xl.ws['B11'].value, '7')

    def test_sort_and_replace(self):
        """
        Test sort and replace using data in column B as sorting keys, 
        and test cell data returns as expected.
        """
        self.xl.sort_and_replace('B', startrow=2)
        self.assertEqual(self.xl.ws['B1'].value, 'Strings')
        self.assertEqual(self.xl.ws['B2'].value, 'Blue')
        self.assertEqual(self.xl.ws['E6'].value, 25.5)
        self.assertEqual(self.xl.ws['C20'].value, 500)

    def test_name_headers(self):
        """
        Test name headers, and test cell data returns as expected.
        """
        self.xl.name_headers({'A': 'TestA', 'D': 'TestD'}, bold=True)
        self.assertEqual(self.xl.ws['A1'].value, 'TestA')
        self.assertEqual(self.xl.ws['B1'].value, 'Strings')
        self.assertEqual(self.xl.ws['D1'].value, 'TestD')
        self.assertEqual(self.xl.ws['B2'].value, 'Red')
        self.iterate_integers()

    def test_get_matching_value(self):
        """
        Test that the value returned from get_matching_value method is
        as expected.
        """
        self.assertEqual(self.xl.get_matching_value('a', 'M', 'B'), 'SNES')
        self.assertEqual(self.xl.get_matching_value('B', 'DS', 'E'), 433.0498)

    def test_set_matching_value(self):
        """
        Uses the set_matching_value method and then checks that the 
        data from the cell returns as expected.
        """
        self.xl.set_matching_value('a', 'O', 'E', 'TEST', startrow=4)
        self.assertEqual(self.xl.ws['E16'].value, 'TEST')

    def test_find_remove_row(self):
        """
        Remove a row based on value 'Switch', and check that a matching
        value from another cell in the same row is not there and that a
        value from the following row is now there in it's place.
        """
        self.xl.find_remove_row('b', 'Switch', startrow=1)
        self.assertFalse(self.xl.ws['E19'].value == 29.5)
        self.assertEqual(self.xl.ws['E19'].value, 433.0498)

    def test_find_replace(self):
        """
        Find and replace the value 'NES' with 'TEST', and then verify
        the cell data returns the new value as expected.
        """
        self.xl.find_replace('B', {'NES': 'TEST'}, ('AB', 'CD'), startrow=2)
        self.assertEqual(self.xl.ws['B13'].value, 'TEST')

    def test_move_values(self):
        """
        Test move_values on pair of columns using a tuple of strings to
        look for, then verifies the cell data returns from the target
        column as expected.
        """
        self.xl.move_values('B', 'C', ('Gameboy', 'Red'))
        self.assertEqual(self.xl.ws['C15'].value, 'Gameboy')

    def test_verify_length(self):
        """
        Test that verify_length runs on an entire column of string
        values. Currently no assert methods.
        """
        self.xl.verify_length('B', 4, 'red', startrow=2)
        self.xl.verify_length('E', 5,
                              'green', skip=['test1', 'test2'], stoprow=8)

    def test_find_and_highlight_rows(self):
        """
        Test that find_and_highlight_rows runs on a specified column. 
        Currently no assert methods.
        """
        self.xl.find_and_highlight_rows('B', 'Cyan', 'yellow', startrow=2)
        self.xl.find_and_highlight_rows('A', 'Q', 'red', startrow=2)
        # TODO self.xl.highlight_rows('c', 300, 'green', startrow=2)
        self.xl.find_and_highlight_rows('d', '05/12/20', 'orange', startrow=2)

    def test_number_type_fix(self):
        """
        Test that number_type_fix runs on all rows of the specified 
        column. Currently no assert methods. 
        """
        self.xl.number_type_fix('C', 'i', startrow=2)
        self.iterate_integers()

    def test_format_date(self):
        """
        Checks value in specified cell is of type:datetime.datetime, 
        runs format_date on a column of data, verifies cell data value 
        returns new value as expected, then checks final value to verify
        it is of type:str.
        """
        self.assertIsInstance(self.xl.ws['D7'].value, datetime.datetime)
        self.xl.format_date('d', startrow=2)
        self.assertEqual(self.xl.ws['D7'].value, '04/25/2019')
        self.assertIsInstance(self.xl.ws['D7'].value, str)

    def test_format_currency(self):
        """
        Test format_currency runs on an entire column, and then verify 
        cell value data is of type:float.
        """
        self.xl.format_currency('E', startrow=2, stoprow=19)
        self.assertIsInstance(self.xl.ws['E15'].value, float)
        self.assertEqual(self.xl.ws['E19'].value, 29.50)

    def test_set_cell_size(self):
        """
        Test set_cell_size runs on a column width as well as a row
        height. Currently no assert methods.
        """
        self.xl.set_cell_size({'A': 48, 2: 20})

    def test_save(self):
        """
        Test save method. Verify "outfile.xlsx" does not exist, save a
        copy of output file using method to be tested, verify the file 
        now does exist, then remove the file again.
        """
        self.assertFalse(Path(f'{tests_path / "outfile.xlsx"}').exists())
        self.xl.save(f'{tests_path / "outfile.xlsx"}')
        self.assertTrue(Path(f'{tests_path / "outfile.xlsx"}').exists())
        Path(f'{tests_path / "outfile.xlsx"}').unlink()
        self.assertFalse(Path(f'{tests_path / "outfile.xlsx"}').exists())

    def test_generate_dictionary(self):
        """
        Test generate_dictionary on main file, then verify that 
        dictionary values match expected data from original file.
        """
        _dict = self.xl.generate_dictionary(
            ('A', 'b', 'd', 'C', 'e'), keycol='b', hdrrow=1)
        self.assertEqual(_dict['Cyan']['Currency'], 48398.58)
        self.assertEqual(_dict['Wii U']['Integers'], 1700)
        self.assertEqual(_dict['DS']['Strings'], 'DS')

    def test_generate_list(self):
        """
        Test generate_list on main file, then verify that list values
        match expected data from original file. Also verify that trying
        to access a list index not included after stop row raises an 
        IndexError.
        """
        _list = self.xl.generate_list(startrow=4, stoprow=14)
        self.assertEqual(_list[0][1], 'Purple')
        self.assertEqual(_list[6][2], 900)
        with self.assertRaises(IndexError):
            _list[11][0]

    def iterate_integers(self, i=100):
        """
        Iterate through Integers row and check for equality. For use in 
        other test methods.
        """
        for row, cell in enumerate(self.xl.ws['C'], 1):
            if row > 1:
                self.assertEqual(cell.value, i)
                i += 100

    def test_search_matching_value(self):
        """
        Tests search_matching_value to verify that the return values
        are as expected.
        """
        check_1 = self.xl.search_matching_value('Strings', 'L')
        self.assertEqual(check_1, 'NES')
        check_2 = self.xl.search_matching_value('Integers', 'S')
        self.assertEqual(check_2, '1800')
        check_3 = self.xl.search_matching_value('Currency', 'Square')
        self.assertEqual(check_3, '20.5')

    def test_set_bold_rows(self):
        """
        Tests set_bold_rows with and without arguments just to make sure
        it runs without errors.
        """
        self.xl.set_bold_rows()
        self.xl.set_bold_rows(startrow=4, stoprow=16)
        self.xl.set_bold_rows(startrow=6)
        self.xl.set_bold_rows(stoprow=18)

    def test_highlight_rows(self):
        """
        Tests highlight_rows with and without arguments to make sure it runs
        with no errors.
        """
        self.xl.highlight_rows()
        self.xl.highlight_rows(startrow=2, stoprow=8,
                               fillcolor='red', alternate=True)
        self.xl.highlight_rows(stoprow=5, alternate=True)
        self.xl.highlight_rows(startrow=3, stoprow=14, fillcolor='yellow')
        self.xl.highlight_rows(fillcolor='BREAK!')


if __name__ == '__main__':
    unittest.main()
